import pyodbc
import pandas as pd
from math import floor
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Database connection details
server = '192.168.29.237'
database = 'ForeThinks'
username = 'medha'
password = 'medha@feb2023'

# Connect to the database
conn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER=' + server + ';DATABASE=' + database + ';UID=' + username + ';PWD=' + password)

# Query the database for unique Put_TIMESTAMP values
timestamps = pd.read_sql("SELECT DISTINCT Put_TIMESTAMP FROM BNF_RawCommon", conn)

# Loop through each unique Put_TIMESTAMP value
for index, row in timestamps.iterrows():
    put_timestamp = row['Put_TIMESTAMP']
    put_timestamp_str = put_timestamp.strftime('%Y_%m_%d')
    
    # Create a new Excel file for this Put_TIMESTAMP value
    file_name = f'D:\\AnlaysisSheets\\Data_{put_timestamp_str}.xlsx'
    wb = Workbook()
    
    # Query the database for unique Put_EXPIRY_DT values for this Put_TIMESTAMP value
    expiry_dates = pd.read_sql(f"SELECT DISTINCT Put_EXPIRY_DT FROM BNF_RawCommon WHERE Put_TIMESTAMP = '{put_timestamp}'", conn)
    
    # Loop through each unique Put_EXPIRY_DT value
    for index, row in expiry_dates.iterrows():
        put_expiry_dt = row['Put_EXPIRY_DT']
        put_expiry_dt_str = put_expiry_dt.strftime('%Y_%m_%d')
        
        # Query the database for data for this Put_TIMESTAMP and Put_EXPIRY_DT value
        data = pd.read_sql(f"SELECT * FROM BNF_RawCommon WHERE Put_TIMESTAMP = '{put_timestamp}' AND Put_EXPIRY_DT = '{put_expiry_dt}'", conn)
        
        # Write the data to a new sheet in the Excel file
        sheet_name = f'common_{put_expiry_dt_str}'
        ws = wb.create_sheet(sheet_name)
        for r in dataframe_to_rows(data, index=False, header=True):
            ws.append(r)
    
    # Query the BNF_Index table for the OPEN value for this Put_TIMESTAMP value
    open_value = pd.read_sql(f"SELECT [OPEN] FROM BNF_Index WHERE HistoricalDate = '{put_timestamp}'", conn).iloc[0]['OPEN']
    
    # Calculate the range of values to search for
    range_start = floor(open_value / 100) * 100
    range_end = range_start + 100
    
    # Loop through each unique Put_EXPIRY_DT value again
    for index, row in expiry_dates.iterrows():
        put_expiry_dt = row['Put_EXPIRY_DT']
        put_expiry_dt_str = put_expiry_dt.strftime('%Y_%m_%d')
        
        # Query the database for data that falls within the specified range for this Put_TIMESTAMP and Put_EXPIRY_DT value
        data = pd.read_sql(f"""
            SELECT * FROM BNF_RawCommon WHERE Put_TIMESTAMP = '{put_timestamp}' AND Put_EXPIRY_DT = '{put_expiry_dt}' AND (
                CMPP_AVG BETWEEN {range_start} AND {range_end} OR 
                CPPM_AVG BETWEEN {range_start} AND {range_end} OR 
                CPPP_AVG BETWEEN {range_start} AND {range_end} OR 
                CMPM_AVG BETWEEN {range_start} AND {range_end}
            )
        """, conn)
        
        # Write the data to a new sheet in the Excel file
        sheet_name = f'OPEN_AVG_{put_expiry_dt_str}'
        ws = wb.create_sheet(sheet_name)
        for r in dataframe_to_rows(data, index=False, header=True):
            ws.append(r)
    
    # Save and close the Excel file
    wb.save(file_name)